import sys
import sqlite3
import pandas as pd
from fpdf import FPDF
from PyQt5.QtWidgets import (
    QHeaderView, QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem, QMessageBox, QComboBox, QDateEdit, QFileDialog
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QIcon, QFont
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import matplotlib.dates as mdates
import openpyxl

# Configuração do banco de dados
conn = sqlite3.connect('financas.db')
cursor = conn.cursor()
cursor.execute('''
    CREATE TABLE IF NOT EXISTS transacoes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        descricao TEXT,
        valor REAL,
        tipo TEXT,
        categoria TEXT,
        data TEXT
    )
''')
conn.commit()

class FinanceiroApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Agente Financeiro Inteligente')
        self.setGeometry(100, 100, 1000, 700)
        self.setWindowIcon(QIcon('logo.png'))

        # Layout principal
        self.showMaximized() 
        main_widget = QWidget()
        main_layout = QVBoxLayout()

        # Formulário de entrada
        form_layout = QHBoxLayout()
        self.descricao_input = QLineEdit()
        self.descricao_input.setPlaceholderText('Descrição')
        self.descricao_input.setStyleSheet("""
            QLineEdit {
                border: 1px solid #6d6d6d;
                border-radius: 5px;
                padding: 2px;
                font-size: 11px;
                background-color: #f9f9f9;
            }
            QLineEdit:focus {
                border: 1px solid #4CAF50;
                background-color: #ffffff;
            }
        """)
        self.valor_input = QLineEdit()
        self.valor_input.setPlaceholderText('Valor')
        self.valor_input.setStyleSheet("""
            QLineEdit {
                border: 1px solid #6d6d6d;
                border-radius: 5px;
                padding: 2px;
                font-size: 11px;
                background-color: #f9f9f9;
            }
            QLineEdit:focus {
                border: 1px solid #4CAF50;
                background-color: #ffffff;
            }
        """)
        self.tipo_input = QComboBox()
        self.tipo_input.addItems(['Receita', 'Despesa'])
        self.tipo_input.setStyleSheet("""
            QComboBox {
                border: 1px solid #6d6d6d;
                border-radius: 5px;
                padding: 2px;
                font-size: 11px;
                background-color: #e2e2e2;
            }
            QComboBox:focus {
                border: 1px solid #4CAF50;
                background-color: #e2e2e2;
            }
            QComboBox::drop-down {
                border: none;
            }
            QComboBox::down-arrow {
                image: url(down-arrow.png);
                width: 8px;
                height: 8px;
            }
        """)
        self.categoria_input = QLineEdit()
        self.categoria_input.setPlaceholderText('Categoria')
        self.categoria_input.setStyleSheet("""
            QLineEdit {
                border: 1px solid #6d6d6d;
                border-radius: 5px;
                padding: 2px;
                font-size: 11px;
                background-color: #f9f9f9;
            }
            QLineEdit:focus {
                border: 1px solid #4CAF50;
                background-color: #ffffff;
            }
        """)
        self.data_input = QDateEdit(calendarPopup=True)
        self.data_input.setDate(QDate.currentDate())
        self.data_input.setStyleSheet("""
            QDateEdit {
                border: 1px solid #6d6d6d;
                border-radius: 5px;
                padding: 2px;
                font-size: 11px;
                background-color: #f9f9f9;
            }
            QDateEdit:focus {
                border: 1px solid #4CAF50;
                background-color: #ffffff;
            }
            QDateEdit::drop-down {
                border: none;
            }
            QDateEdit::down-arrow {
                image: url(down-arrow.png);
                width: 8px;
                height: 8px;
            }
        """)
        adicionar_button = QPushButton('Adicionar')
        adicionar_button.clicked.connect(self.adicionar_transacao)

        form_layout.addWidget(self.descricao_input)
        form_layout.addWidget(self.valor_input)
        form_layout.addWidget(self.tipo_input)
        form_layout.addWidget(self.categoria_input)
        form_layout.addWidget(self.data_input)
        form_layout.addWidget(adicionar_button)

        # Tabela de transações
        self.tabela = QTableWidget()
        self.tabela.setColumnCount(6)
        self.tabela.setHorizontalHeaderLabels(['ID', 'Descrição', 'Valor', 'Tipo', 'Categoria', 'Data'])
        self.tabela.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tabela.setSelectionBehavior(QTableWidget.SelectRows)
        self.tabela.setStyleSheet('border-radius: 5px;')
        self.tabela.setFont(QFont('Arial', 9))
        self.tabela.horizontalHeader().setStretchLastSection(True)
        self.tabela.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        # Botões de controle
        controle_layout = QHBoxLayout()
        atualizar_button = QPushButton('Atualizar')
        atualizar_button.clicked.connect(self.carregar_transacoes)
        resumo_button = QPushButton('Resumo')
        resumo_button.clicked.connect(self.mostrar_resumo)
        apagar_button = QPushButton('Apagar')
        apagar_button.clicked.connect(self.apagar_transacao)
        exportar_excel_button = QPushButton('Exportar para Excel')
        exportar_excel_button.clicked.connect(self.exportar_excel)
        exportar_pdf_button = QPushButton('Exportar para PDF')
        exportar_pdf_button.clicked.connect(self.exportar_pdf)

        controle_layout.addWidget(atualizar_button)
        controle_layout.addWidget(resumo_button)
        controle_layout.addWidget(apagar_button)
        controle_layout.addWidget(exportar_excel_button)
        controle_layout.addWidget(exportar_pdf_button)

        # Filtros de data
        filtro_layout = QHBoxLayout()
        self.data_inicio = QDateEdit(calendarPopup=True)
        self.data_inicio.setDate(QDate.currentDate().addMonths(-1))
        self.data_inicio.setStyleSheet("""
            QDateEdit {
                border: 1px solid #6d6d6d;
                border-radius: 5px;
                padding: 2px;
                font-size: 11px;
                background-color: #f9f9f9;
            }
            QDateEdit:focus {
                border: 1px solid #4CAF50;
                background-color: #ffffff;
            }
            QDateEdit::drop-down {
                border: none;
            }
            QDateEdit::down-arrow {
                image: url(down-arrow.png);
                width: 8px;
                height: 8px;
            }
        """)
        self.data_fim = QDateEdit(calendarPopup=True)
        self.data_fim.setDate(QDate.currentDate())
        self.data_fim.setStyleSheet("""
            QDateEdit {
                border: 1px solid #6d6d6d;
                border-radius: 5px;
                padding: 2px;
                font-size: 11px;
                background-color: #f9f9f9;
            }
            QDateEdit:focus {
                border: 1px solid #4CAF50;
                background-color: #ffffff;
            }
            QDateEdit::drop-down {
                border: none;
            }
            QDateEdit::down-arrow {
                image: url(down-arrow.png);
                width: 8px;
                height: 8px;
            }
        """)
        filtrar_button = QPushButton('Filtrar')
        filtrar_button.clicked.connect(self.filtrar_transacoes)

        filtro_layout.addWidget(QLabel('Data Início:'))
        filtro_layout.addWidget(self.data_inicio)
        filtro_layout.addWidget(QLabel('Data Fim:'))
        filtro_layout.addWidget(self.data_fim)
        filtro_layout.addWidget(filtrar_button)

        # Gráfico de resumo
        self.figura = Figure()
        self.canvas = FigureCanvas(self.figura)

        # Adicionando layouts ao layout principal
        main_layout.addLayout(form_layout)
        main_layout.addLayout(filtro_layout)
        main_layout.addWidget(self.tabela)
        main_layout.addLayout(controle_layout)
        main_layout.addWidget(self.canvas)

        main_widget.setLayout(main_layout)
        self.setCentralWidget(main_widget)

        # Estilizando os botões
        self.estilizar_botoes({
            adicionar_button: '#4CAF50',  # Verde
            atualizar_button: '#2196F3',  # Azul
            resumo_button: '#ffbb2f',    # Amarelo
            apagar_button: '#F44336',    # Vermelho
            exportar_excel_button: '#8BC34A',  # Verde claro
            exportar_pdf_button: '#FF5722',    # Laranja
            filtrar_button: '#A2A2A2'    # Cinza
        })

        # Carregar transações ao iniciar
        self.carregar_transacoes()

    def estilizar_botoes(self, botoes_cores):
        estilo_base = """
            QPushButton {{
                background-color: {cor};
                color: white;
                border: none;
                border-radius: 5px;
                padding: 4px 10px;
                font-size: 12px;
            }}
            QPushButton:hover {{
                background-color: {cor_hover};
            }}
            QPushButton:pressed {{
                background-color: {cor_pressed};
            }}
        """
        for botao, cor in botoes_cores.items():
            cor_hover = self.escurecer_cor(cor, 0.1)
            cor_pressed = self.escurecer_cor(cor, 0.2)
            estilo = estilo_base.format(cor=cor, cor_hover=cor_hover, cor_pressed=cor_pressed)
            botao.setStyleSheet(estilo)

    def escurecer_cor(self, cor, fator):
        cor = cor.lstrip('#')
        r = int(cor[0:2], 16)
        g = int(cor[2:4], 16)
        b = int(cor[4:6], 16)

        r = max(0, min(255, int(r * (1 - fator))))
        g = max(0, min(255, int(g * (1 - fator))))
        b = max(0, min(255, int(b * (1 - fator))))

        return f'#{r:02x}{g:02x}{b:02x}'

    def adicionar_transacao(self):
        descricao = self.descricao_input.text()
        valor = self.valor_input.text()
        tipo = self.tipo_input.currentText()
        categoria = self.categoria_input.text()
        data = self.data_input.date().toString('yyyy-MM-dd')

        if not descricao or not valor or not categoria:
            QMessageBox.warning(self, 'Erro', 'Todos os campos devem ser preenchidos!')
            return

        try:
            valor = float(valor)
        except ValueError:
            QMessageBox.warning(self, 'Erro', 'Valor deve ser um número!')
            return

        cursor.execute('''
            INSERT INTO transacoes (descricao, valor, tipo, categoria, data)
            VALUES (?, ?, ?, ?, ?)
        ''', (descricao, valor, tipo, categoria, data))
        conn.commit()

        self.descricao_input.clear()
        self.valor_input.clear()
        self.categoria_input.clear()

        self.carregar_transacoes()

    def carregar_transacoes(self):
        cursor.execute('SELECT * FROM transacoes')
        transacoes = cursor.fetchall()

        self.tabela.setRowCount(len(transacoes))
        for row, transacao in enumerate(transacoes):
            for col, item in enumerate(transacao):
                self.tabela.setItem(row, col, QTableWidgetItem(str(item)))

    def filtrar_transacoes(self):
        data_inicio = self.data_inicio.date().toString('yyyy-MM-dd')
        data_fim = self.data_fim.date().toString('yyyy-MM-dd')

        cursor.execute('''
            SELECT * FROM transacoes
            WHERE data BETWEEN ? AND ?
        ''', (data_inicio, data_fim))
        transacoes = cursor.fetchall()

        self.tabela.setRowCount(len(transacoes))
        for row, transacao in enumerate(transacoes):
            for col, item in enumerate(transacao):
                self.tabela.setItem(row, col, QTableWidgetItem(str(item)))

    def apagar_transacao(self):
        selected_rows = self.tabela.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, 'Erro', 'Nenhuma transação selecionada!')
            return

        for index in selected_rows:
            transacao_id = self.tabela.item(index.row(), 0).text()
            cursor.execute('DELETE FROM transacoes WHERE id = ?', (transacao_id,))
            conn.commit()

        self.carregar_transacoes()

    def mostrar_resumo(self):
        cursor.execute('SELECT SUM(valor) FROM transacoes WHERE tipo = "Receita"')
        total_receitas = cursor.fetchone()[0] or 0

        cursor.execute('SELECT SUM(valor) FROM transacoes WHERE tipo = "Despesa"')
        total_despesas = cursor.fetchone()[0] or 0

        saldo = total_receitas - total_despesas

        QMessageBox.information(self, 'Resumo Financeiro', f'''
            Total de Receitas: R$ {total_receitas:.2f}
            Total de Despesas: R$ {total_despesas:.2f}
            Saldo: R$ {saldo:.2f}
        ''')

        # Gráfico de resumo para Despesas por Categoria
        cursor.execute('''
            SELECT categoria, SUM(valor) FROM transacoes
            WHERE tipo = "Despesa"
            GROUP BY categoria
        ''')
        despesas_por_categoria = cursor.fetchall()

        categorias_despesas = [item[0] for item in despesas_por_categoria]
        valores_despesas = [item[1] for item in despesas_por_categoria]

        # Gráfico de resumo para Receitas por Categoria
        cursor.execute('''
            SELECT categoria, SUM(valor) FROM transacoes
            WHERE tipo = "Receita"
            GROUP BY categoria
        ''')
        receitas_por_categoria = cursor.fetchall()

        categorias_receitas = [item[0] for item in receitas_por_categoria]
        valores_receitas = [item[1] for item in receitas_por_categoria]

        self.figura.clear()
        ax1 = self.figura.add_subplot(223)
        ax1.pie(valores_despesas, labels=categorias_despesas, autopct='%1.1f%%', startangle=140)
        ax1.set_title('Despesas por Categoria')

        ax2 = self.figura.add_subplot(221)
        ax2.pie(valores_receitas, labels=categorias_receitas, autopct='%1.1f%%', startangle=140)
        ax2.set_title('Receitas por Categoria')

        # Gráfico de colunas para receitas e despesas ao longo do tempo
        cursor.execute('''
            SELECT strftime('%m/%Y', data) as Mes, SUM(valor) as Total
            FROM transacoes
            WHERE tipo = "Receita"
            GROUP BY Mes
            ORDER BY Mes
        ''')
        receitas_por_mes = cursor.fetchall()

        cursor.execute('''
            SELECT strftime('%m/%Y', data) as Mes, SUM(valor) as Total
            FROM transacoes
            WHERE tipo = "Despesa"
            GROUP BY Mes
            ORDER BY Mes
        ''')
        despesas_por_mes = cursor.fetchall()

        # Preparando os dados para o gráfico
        meses = sorted(set([item[0] for item in receitas_por_mes] + [item[0] for item in despesas_por_mes]))
        receitas_dict = {item[0]: item[1] for item in receitas_por_mes}
        despesas_dict = {item[0]: item[1] for item in despesas_por_mes}

        receitas_vals = [receitas_dict.get(mes, 0) for mes in meses]
        despesas_vals = [despesas_dict.get(mes, 0) for mes in meses]

        ax3 = self.figura.add_subplot(122)
        bar_width = 0.35
        index = range(len(meses))

        ax3.bar(index, receitas_vals, bar_width, label='Receitas', alpha=0.6, color='blue')
        ax3.bar([i + bar_width for i in index], despesas_vals, bar_width, label='Despesas', alpha=0.6, color='red')

        ax3.set_xlabel('Mês')
        ax3.set_ylabel('Valor (R$)')
        ax3.set_title('Receitas e Despesas por Mês')
        ax3.set_xticks([i + bar_width / 2 for i in index])
        ax3.set_xticklabels(meses)
        ax3.legend()
        ax3.grid(True)

        self.canvas.draw()

    def exportar_excel(self):
        file_path, _ = QFileDialog.getSaveFileName(self, 'Salvar Relatório', '', 'Excel Files (*.xlsx)')
        if not file_path:
            return

        cursor.execute('SELECT * FROM transacoes')
        transacoes = cursor.fetchall()

        df = pd.DataFrame(transacoes, columns=['ID', 'Descrição', 'Valor', 'Tipo', 'Categoria', 'Data'])
        df.to_excel(file_path, index=False)

        QMessageBox.information(self, 'Sucesso', 'Relatório exportado com sucesso!')

    def exportar_pdf(self):
        file_path, _ = QFileDialog.getSaveFileName(self, 'Salvar Relatório', '', 'PDF Files (*.pdf)')
        if not file_path:
            return

        cursor.execute('SELECT * FROM transacoes')
        transacoes = cursor.fetchall()

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 10, 'Relatório Financeiro', ln=True, align='C')
        pdf.ln(10)

        pdf.set_font('Arial', 'B', 12)
        pdf.cell(20, 10, 'ID', 1)
        pdf.cell(60, 10, 'Descrição', 1)
        pdf.cell(30, 10, 'Valor', 1)
        pdf.cell(30, 10, 'Tipo', 1)
        pdf.cell(30, 10, 'Categoria', 1)
        pdf.cell(30, 10, 'Data', 1)
        pdf.ln()

        pdf.set_font('Arial', '', 12)
        for transacao in transacoes:
            pdf.cell(20, 10, str(transacao[0]), 1)
            pdf.cell(60, 10, transacao[1], 1)
            pdf.cell(30, 10, f'R$ {transacao[2]:.2f}', 1)
            pdf.cell(30, 10, transacao[3], 1)
            pdf.cell(30, 10, transacao[4], 1)
            pdf.cell(30, 10, transacao[5], 1)
            pdf.ln()

        pdf.output(file_path)

        QMessageBox.information(self, 'Sucesso', 'Relatório exportado com sucesso!')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = FinanceiroApp()
    window.show()
    sys.exit(app.exec_())
